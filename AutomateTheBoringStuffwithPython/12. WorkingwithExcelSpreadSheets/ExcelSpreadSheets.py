import openpyxl
import os

def basicSheetsCommand(wb):

    print(type(wb))
    print(wb.sheetnames)
    sheet1 = wb['Sheet1']
    print(sheet1.title)
    print(sheet1['B1'].value)
    print(sheet1.cell(row=1, column=2).value)

def runOverMultipleRows(wb):
    sheet1 = wb['Sheet1']

    # The merged cell will put the value in the top left corner
    for rowsCellObject in sheet1['V9':'W10']: # returning a tuple object
        print(rowsCellObject)
        for cellObj in rowsCellObject: # iterate over a tuple cell by cell
            if cellObj.value !=None:
                print(cellObj.coordinate, cellObj.value)
        print('END OF ROW')


    for rowsCellObject in sheet1['D4':'R4']: # returning a tuple object
        for cellObj in rowsCellObject:
            print(cellObj.value)

def mergingExample(filename,wb):
    sheet1 = wb['MySheet']
    sheet1.merge_cells('A1:D3')
    sheet1['A1'] = 'Twelve cells merged together.'
    wb.save(filename)

def writingAndSaving(wb):

    sheet1 = wb['Sheet1']
    sheet1['A1'] = 'Written from Python'

    counter = 1
    base_filename, extension = os.path.splitext("MaHaLuzExample.xlsx")
    new_filename = f"{base_filename}_{counter}{extension}"

    while os.path.exists(new_filename):
        counter += 1
        new_filename = f"{base_filename}_{counter}{extension}"
    wb.save(new_filename)


def main():
    file_name = 'MaHaLuzExample.xlsx'
    wb = openpyxl.load_workbook(file_name)
    #basicSheetsCommand(wb)
    #runOverMultipleRows(wb)
    #mergingExample(file_name,wb)
    #writingAndSaving(wb)

if __name__ == '__main__':
    main()