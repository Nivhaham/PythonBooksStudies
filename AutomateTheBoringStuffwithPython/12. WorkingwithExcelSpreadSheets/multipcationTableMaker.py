import openpyxl
import sys
import os


def makeExcel():
    wb = openpyxl.Workbook()
    file_name = "MultiplicationTableExcel.xlsx"

    if not os.path.exists(file_name):
        wb.save(file_name)
        print(f"Excel file '{file_name}' created successfully.")
    else:
        print(f"'{file_name}' Already Exists File")
    return file_name


def nameSheet(file_name, wb, name):
    ws = wb.active
    ws.title = name
    wb.save(file_name)


def buildRowAndColumn(file_name, wb):
    if len(sys.argv) < 2:
        N = 5
    else:
        N = sys.argv[1]
    print(N)
    N = int(N)
    start = 'B'

    for i in range(2, N + 2):
        wb.active['A' + str(i)] = i-1
        wb.active[start + '1'] = i-1
        start = chr(ord(start) + 1)

    wb.save(file_name)
    return N

def calculateMatrix(file_name, wb,N):
    ws = wb.active
    start = 'B'
    for i in range(2, N + 2):
        for j in range(0, N):
            currChar = chr(ord(start) + j)
            ws[currChar + str(i)] = f"={currChar}1*A{str(i)}"
    wb.save(file_name)


def main():
    file_name = makeExcel()
    wb = openpyxl.load_workbook(file_name)
    # name = "Multipcation"
    # nameSheet(file_name, wb, name)
    N = buildRowAndColumn(file_name, wb)
    calculateMatrix(file_name, wb, N)


if __name__ == '__main__':
    main()
